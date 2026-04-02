"""
Support Team Email Extractor
Scans the shared mailbox "Support Team" in Outlook and exports emails to Excel.

SETUP:
  1. Create a .env file with:
       SHARED_MAILBOX=supportteam@auckland.ac.nz
  2. Make sure Outlook is open and the shared mailbox is visible in the sidebar
  3. Adjust START_DATE, END_DATE, and TARGET_FOLDERS below
  4. Run:  python support_email_extractor.py
"""

import os
import re
import datetime
import win32com.client
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# ===================== CONFIG =====================

SHARED_MAILBOX = os.getenv("SHARED_MAILBOX", "supportteam@auckland.ac.nz")

# Date range — local NZ time, no UTC conversion needed
START_DATE = datetime.datetime(2026, 3, 25, 0, 0, 0)
END_DATE   = datetime.datetime(2026, 3, 31, 23, 59, 59)

# Which folders to scan inside the shared mailbox.
# Set to None to scan the top-level Inbox + all its subfolders.
# Or list specific subfolder names like: ["Agent emails/INTL office", "Done"]
TARGET_FOLDERS = None  # None = scan everything under Inbox


# Output
OUTPUT_FOLDER = os.path.join(os.environ.get("USERPROFILE", "."), "Documents", "student-applications", "support-tracker")
OUTPUT_FILE   = f"support_tracker_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"

# Signature triggers — lines matching these get stripped (and everything below them)
SIGNATURE_TRIGGERS = [
    "ngā mihi", "nga mihi", "kind regards", "regards", "thanks,",
    "thank you,", "cheers,", "ngā mihi maioha",
]

# ===================== END CONFIG =====================


def find_shared_mailbox(outlook_ns, shared_email):
    """Find the shared mailbox root folder in Outlook."""

    # Method 1: Walk Stores (mailbox fully added to profile)
    for store in outlook_ns.Stores:
        try:
            display = (store.DisplayName or "").lower()
            if shared_email.lower() in display or "support team" in display:
                return store.GetRootFolder()
        except Exception:
            continue

    # Method 2: GetSharedDefaultFolder (delegate access)
    try:
        recipient = outlook_ns.CreateRecipient(shared_email)
        recipient.Resolve()
        if recipient.Resolved:
            inbox = outlook_ns.GetSharedDefaultFolder(recipient, 6)
            return inbox.Parent  # return root so we can also get Sent Items
    except Exception:
        pass

    return None


def get_all_folders(parent_folder, depth=0, max_depth=5):
    """Recursively get all mail folders under a parent."""
    folders = []
    try:
        for folder in parent_folder.Folders:
            folders.append((folder, "  " * depth + folder.Name))
            if depth < max_depth:
                folders.extend(get_all_folders(folder, depth + 1, max_depth))
    except Exception:
        pass
    return folders


def strip_signature(body):
    """Remove signature block from email body."""
    lines = body.splitlines()
    clean = []
    for line in lines:
        stripped = line.strip().lower().rstrip(",")
        if any(stripped == t.rstrip(",") for t in SIGNATURE_TRIGGERS):
            break
        clean.append(line)

    while clean and not clean[-1].strip():
        clean.pop()
    return "\n".join(clean)


def strip_html(html):
    """Basic HTML to plain text."""
    text = re.sub(r"<(style|script)[^>]*>.*?</(style|script)>", "", html, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<(br|p|div|tr|li)[^>]*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    for entity, char in [("&amp;","&"),("&lt;","<"),("&gt;",">"),("&nbsp;"," "),("&#39;","'"),("&quot;",'"')]:
        text = text.replace(entity, char)
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def fetch_emails_from_folder(folder, folder_name):
    """Fetch all emails from a single folder within the date range."""
    emails = []

    try:
        items = folder.Items
        items.Sort("[ReceivedTime]", True)

        # Use local time — no UTC conversion
        start_str = START_DATE.strftime("%m/%d/%Y %I:%M %p")
        end_str   = END_DATE.strftime("%m/%d/%Y %I:%M %p")
        restrict  = f"[ReceivedTime] >= '{start_str}' AND [ReceivedTime] <= '{end_str}'"

        filtered = items.Restrict(restrict)
        count = filtered.Count
        print(f"   📊 {count} items in date range")

        for msg in filtered:
            try:
                rt = msg.ReceivedTime
                msg_dt = datetime.datetime(rt.year, rt.month, rt.day, rt.hour, rt.minute, rt.second)

                # Double-check date range (belt and suspenders)
                if not (START_DATE <= msg_dt <= END_DATE):
                    continue

                subject = msg.Subject or ""
                sender  = msg.SenderName or ""

                # Get body — prefer plain text, fall back to HTML
                try:
                    body = msg.Body or ""
                    if not body.strip():
                        body = strip_html(msg.HTMLBody or "")
                except Exception:
                    body = ""

                body_clean = strip_signature(body)

                # Try to get conversation ID for reply matching
                try:
                    conv_id = msg.ConversationID or ""
                except Exception:
                    conv_id = ""

                emails.append({
                    "received":    msg_dt,
                    "subject":     subject,
                    "sender":      sender,
                    "body":        body_clean,
                    "folder":      folder_name,
                    "conv_id":     conv_id,
                })

            except Exception as e:
                print(f"   ⚠️  Skipped one email: {e}")
                continue

    except Exception as e:
        print(f"   ❌ Error scanning folder: {e}")

    return emails


def fetch_sent_replies(shared_root):
    """Build a lookup of ConversationID -> reply info from Sent Items."""
    sent_map = {}

    sent_folder = None
    try:
        for folder in shared_root.Folders:
            if folder.Name.lower() == "sent items":
                sent_folder = folder
                break
    except Exception:
        return sent_map

    if not sent_folder:
        print("   ⚠️  No 'Sent Items' folder found — Reply Sent column will be empty")
        return sent_map

    try:
        items = sent_folder.Items
        items.Sort("[SentOn]", True)
        start_str = START_DATE.strftime("%m/%d/%Y %I:%M %p")
        end_str   = END_DATE.strftime("%m/%d/%Y %I:%M %p")
        restrict  = f"[SentOn] >= '{start_str}' AND [SentOn] <= '{end_str}'"
        items = items.Restrict(restrict)

        for msg in items:
            try:
                conv_id = msg.ConversationID or ""
                if conv_id and conv_id not in sent_map:
                    st = msg.SentOn
                    sent_map[conv_id] = datetime.datetime(
                        st.year, st.month, st.day, st.hour, st.minute, st.second
                    ).strftime("%d/%m/%Y %H:%M")
            except Exception:
                continue

        print(f"   ✅ Found {len(sent_map)} sent replies")
    except Exception as e:
        print(f"   ⚠️  Error scanning sent items: {e}")

    return sent_map


def extract_id(subject):
    """Pull ID# number from subject if present."""
    m = re.search(r"id#?\s*(\d{6,12})", subject, re.IGNORECASE)
    return m.group(1) if m else ""


def detect_request_type(subject):
    """Categorise email by subject keywords."""
    s = subject.lower()
    patterns = {
        "Single Name (Temp)": ["single name"],
        "Name Order":         ["name order"],
        "Name Change":        ["name change"],
        "Name Correction":    ["name correction"],
        "DOB Correction":     ["dob correction", "date of birth correction"],
    }
    for rtype, keywords in patterns.items():
        for kw in keywords:
            if kw in s:
                return rtype
    return ""


def write_excel(emails, sent_map, output_path):
    """Write emails to a formatted Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Email Tracker"

    columns = [
        ("ID#",              14),
        ("Request Type",     20),
        ("Folder",           22),
        ("Student Name",     22),
        ("Received Date",    14),
        ("Received Time",    12),
        ("Sent By",          22),
        ("Email Subject",    40),
        ("Email Body",       60),
        ("Reply Sent?",      12),
        ("Reply Sent At",    16),
        ("Notes",            30),
        ("Status",           14),
    ]

    header_bg   = "1F3864"
    alt_bg      = "EEF2F7"
    tick_green   = "E8F5E9"
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        left=Side(style="thin", color="EEEEEE"),
        right=Side(style="thin", color="EEEEEE"),
    )

    # Header row
    for col_idx, (name, width) in enumerate(columns, 1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", start_color=header_bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(bottom=Side(style="medium", color="AAAAAA"))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, email in enumerate(emails, 2):
        reply_at = sent_map.get(email["conv_id"], "")
        reply_yn = "Yes" if reply_at else "No"

        row_data = {
            "ID#":            extract_id(email["subject"]),
            "Request Type":   detect_request_type(email["subject"]),
            "Folder":         email["folder"],
            "Student Name":   "",
            "Received Date":  email["received"].strftime("%d/%m/%Y"),
            "Received Time":  email["received"].strftime("%H:%M"),
            "Sent By":        email["sender"],
            "Email Subject":  email["subject"],
            "Email Body":     email["body"],
            "Reply Sent?":    reply_yn,
            "Reply Sent At":  reply_at,
            "Notes":          "",
            "Status":         "",
        }

        base_fill = PatternFill("solid", start_color=alt_bg if row_idx % 2 == 0 else "FFFFFF")

        for col_idx, (col_name, _) in enumerate(columns, 1):
            c        = ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
            c.font   = Font(name="Arial", size=10)
            c.border = thin_border

            if col_name == "Email Body":
                c.alignment = Alignment(wrap_text=True, vertical="top")
                c.fill = base_fill
                ws.row_dimensions[row_idx].height = 80
            elif col_name == "Email Subject":
                c.alignment = Alignment(wrap_text=True, vertical="center")
                c.fill = base_fill
            else:
                c.alignment = Alignment(vertical="center")
                c.fill = base_fill

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    last_row = len(emails) + 1
    ws2["A1"] = "SUMMARY"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color=header_bg)

    summary_items = [
        ("Total emails", f"=COUNTA('Email Tracker'!A2:A{last_row})"),
        ("", ""),
        ("By Request Type", ""),
        ("Single Name (Temp)", f"=COUNTIF('Email Tracker'!B2:B{last_row},\"Single Name (Temp)\")"),
        ("Name Order",         f"=COUNTIF('Email Tracker'!B2:B{last_row},\"Name Order\")"),
        ("Name Change",        f"=COUNTIF('Email Tracker'!B2:B{last_row},\"Name Change\")"),
        ("Name Correction",    f"=COUNTIF('Email Tracker'!B2:B{last_row},\"Name Correction\")"),
        ("DOB Correction",     f"=COUNTIF('Email Tracker'!B2:B{last_row},\"DOB Correction\")"),
    ]

    for r, (label, val) in enumerate(summary_items, 2):
        ws2.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=(val == ""), size=10)
        ws2.cell(row=r, column=2, value=val).font   = Font(name="Arial", size=10)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    wb.save(output_path)
    print(f"\n✅ Saved: {output_path}")
    print(f"   📊 {len(emails)} rows")


# ===================== MAIN =====================

if __name__ == "__main__":
    print("=" * 60)
    print("  SUPPORT TEAM EMAIL EXTRACTOR")
    print("=" * 60)
    print(f"  Shared mailbox : {SHARED_MAILBOX}")
    print(f"  Date range     : {START_DATE.strftime('%d/%m/%Y')} → {END_DATE.strftime('%d/%m/%Y')}")
    print()

    # Create output folder
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    # Connect to Outlook
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        print("✅ Connected to Outlook")
    except Exception as e:
        print(f"❌ Could not connect to Outlook. Is it open?\n   {e}")
        exit(1)

    # Find shared mailbox
    shared_root = find_shared_mailbox(outlook, SHARED_MAILBOX)
    if not shared_root:
        print(f"❌ Could not find shared mailbox: {SHARED_MAILBOX}")
        print("   Make sure it appears in your Outlook sidebar.")
        exit(1)
    print("✅ Found shared mailbox")

    # List all folders so we know what's there
    # print("\n📁 Folders in shared mailbox:")
    # all_folders = get_all_folders(shared_root)
    # for folder, display_name in all_folders:
    #     try:
    #         count = folder.Items.Count
    #         print(f"   {display_name} ({count})")
    #     except Exception:
    #         print(f"   {display_name}")

    # Find the Inbox
    inbox = None
    for folder in shared_root.Folders:
        if folder.Name.lower() == "inbox":
            inbox = folder
            break

    if not inbox:
        print("❌ Could not find Inbox in shared mailbox")
        exit(1)

    # Decide which folders to scan
    folders_to_scan = []

    folders_to_scan.append((inbox, "Inbox"))

    ROOT_FOLDERS = ["Student Systems/Operation"]
    for folder in shared_root.Folders:
        if folder.Name in ROOT_FOLDERS:
            folders_to_scan.append((folder, folder.Name))
            print(f"   ✅ Added root folder: {folder.Name}")

    # if not folders_to_scan:
    #     print("❌ No folders to scan")
    #     exit(1)

    # Fetch emails from all target folders
    all_emails = []
    for folder, label in folders_to_scan:
        print(f"\n📂 Scanning: {label}")
        folder_emails = fetch_emails_from_folder(folder, label)
        print(f"   ✅ {len(folder_emails)} emails collected")
        all_emails.extend(folder_emails)

    if not all_emails:
        print("\n⚠️  No emails found in the date range.")
        print("   Try widening START_DATE / END_DATE in the config.")
        exit(0)

    # Sort by date (newest first)
    all_emails.sort(key=lambda e: e["received"], reverse=True)
    print(f"\n📧 Total emails collected: {len(all_emails)}")

    # Fetch sent replies
    print("\n📤 Checking Sent Items for replies...")
    sent_map = fetch_sent_replies(shared_root)

    # Write Excel
    output_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILE)
    write_excel(all_emails, sent_map, output_path)

    # Auto-open
    try:
        os.startfile(output_path)
    except Exception:
        pass

    print("\n✅ DONE")